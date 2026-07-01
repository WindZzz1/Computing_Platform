# -*- coding: utf-8 -*-
"""
重新生成 6 个教学班的“成绩导入”Excel（覆盖 excel/4-成绩-SE10x_*.xlsx）。

关键点：
- 表头与数据库考核点配置【完全一致】：AP1 课堂测验 30 / AP2 实验报告 40 / AP3 课程设计 20 / AP4 期末考试 10
  （旧文件的表头名称和满分数值都是错的，且系统导入不读表头、只按列位置匹配数据库考核点，
   导致按旧表头填的分被按真实满分拦截 —— 这就是之前 AP4 报“超过满分 10.0”的根因）
- 每个考核点得分 ≤ 其真实满分，可通过系统“导入成绩”校验
- 5 个学生按能力分 5 档（优/良/中/及格/弱）；6 个班整体偏移不同，便于测试达成度/报表的差异
- 所有数值确定性生成（无随机），可复现、可预期
"""
import os
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

EXCEL_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "excel")

# 6 个班：(SE代码, 课程名, [5个学号·升序])
CLASSES = [
    ("SE101", "软件工程导论", ["202301001", "202301002", "202301003", "202301004", "202301005"]),
    ("SE102", "数据结构与算法", ["202301006", "202301007", "202301008", "202301009", "202301010"]),
    ("SE103", "操作系统",       ["202301011", "202301012", "202301013", "202301014", "202301015"]),
    ("SE104", "计算机网络",     ["202301016", "202301017", "202301018", "202301019", "202301020"]),
    ("SE105", "数据库原理",     ["202301021", "202301022", "202301023", "202301024", "202301025"]),
    ("SE106", "软件测试",       ["202301026", "202301027", "202301028", "202301029", "202301030"]),
]

# 考核点：(名称, 满分) —— 与数据库一致；列顺序 = point_code 升序，匹配导入解析顺序
POINTS = [("课堂测验", 30), ("实验报告", 40), ("课程设计", 20), ("期末考试", 10)]

# 5 档达成率（行=档位 0优..4弱；列=AP1..AP4）
RATIOS = [
    [0.95, 0.90, 0.92, 0.88],   # 优
    [0.85, 0.82, 0.80, 0.78],   # 良
    [0.72, 0.75, 0.70, 0.65],   # 中
    [0.60, 0.62, 0.58, 0.55],   # 及格
    [0.48, 0.50, 0.45, 0.40],   # 弱
]
# 6 个班的整体偏移（确定性），让各班达成度有区分
DELTAS = [0.00, 0.04, -0.04, 0.08, -0.08, 0.02]


def student_name(no):
    return f"测试学生{int(no[-2:]):02d}"


def gen_scores(delta):
    """返回 5×4 得分矩阵，每项 clamp 到 [0, 满分]。"""
    rows = []
    for level in range(5):
        scores = []
        for p in range(4):
            r = max(0.10, min(0.99, RATIOS[level][p] + delta))
            scores.append(round(POINTS[p][1] * r, 1))
        rows.append(scores)
    return rows


def write_excel(se, course, student_nos, score_rows, suffix=""):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "成绩录入"

    headers = ["学号", "姓名"]
    for i, (name, full) in enumerate(POINTS, start=1):
        headers.append(f"AP{i}-{course}-{name}\n满分:{full:.1f}")
    ws.append(headers)

    for no, scores in zip(student_nos, score_rows):
        ws.append([no, student_name(no)] + scores)

    # 样式
    head_fill = PatternFill("solid", fgColor="D9E1F2")
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = head_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for r in range(2, ws.max_row + 1):
        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).alignment = Alignment(horizontal="center", vertical="center")
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 12
    for col in "CDEF":
        ws.column_dimensions[col].width = 20
    ws.row_dimensions[1].height = 34

    path = os.path.join(EXCEL_DIR, f"4-成绩-{se}_{course}{suffix}.xlsx")
    wb.save(path)
    return path


def main():
    ok, fail, warn = [], [], []
    for idx, (se, course, nos) in enumerate(CLASSES):
        scores = gen_scores(DELTAS[idx])
        # 合法性自检：每项 ∈ [0, 满分]
        for row in scores:
            for v, (_, full) in zip(row, POINTS):
                assert 0 <= v <= full, f"{se} {course} 越界: {v} > {full}"
        try:
            write_excel(se, course, nos, scores)
            ok.append((se, course, nos, scores))
        except PermissionError:
            # 原文件被 Excel 占用，写到带 _新 后缀的备选文件，避免阻塞
            try:
                write_excel(se, course, nos, scores, suffix="_新")
                ok.append((se, course, nos, scores))
                warn.append((se, course))
            except Exception as e:
                fail.append((se, course, str(e)))

    print("=== 生成结果（满分: AP1=30  AP2=40  AP3=20  AP4=10）===")
    for se, course, nos, scores in ok:
        print(f"\n[{se}] {course}  班均≈{round(sum(sum(s) for s in scores)/5,1)}")
        for no, s in zip(nos, scores):
            print(f"  {no} {student_name(no)}  AP1={s[0]:>4}  AP2={s[1]:>4}  AP3={s[2]:>4}  AP4={s[3]:>4}   合计={round(sum(s),1):>5}")
    if warn:
        print("\n以下原文件被 Excel 占用，已生成到「_新.xlsx」备选文件（关闭原 Excel 后可覆盖回原名）：")
        for se, course in warn:
            print(f"  [WARN] 4-成绩-{se}_{course}.xlsx → 4-成绩-{se}_{course}_新.xlsx")
    if fail:
        print("\n以下文件生成失败：")
        for item in fail:
            print(f"  [FAIL] {item}")
    print(f"\n成功 {len(ok)} 个 / 失败 {len(fail)} 个")


if __name__ == "__main__":
    main()
