# -*- coding: utf-8 -*-
"""
生成课堂演示用 Excel（面向专业认证的毕业要求达成度统一计算平台）。
数据设计与 sql/generate_test_data.py 完全一致（软件工程专业 / 6 课 / 30 学生 / 12 指标点，权重 ∑=1.0）。
- 1-课程导入.xlsx   → edu 用 /course/import/excel
- 2-学生导入.xlsx   → edu 用 /student/import/excel
- 3-教学班学生-*.xlsx (6个) → edu 用 /teaching-class/import-students/excel（按教学班导入）
- 4-成绩-*.xlsx     (6个) → teacher 用 /grade-entry/import（按教学班导入，列序匹配 AP1..AP4）
字段严格匹配后端 CourseExcel / StudentExcel / ClassStudentExcel / 成绩动态模板。
"""
import os
import openpyxl
from openpyxl.styles import Alignment, Font

OUT = os.path.join(os.path.dirname(__file__), "excel")
os.makedirs(OUT, exist_ok=True)

MAJOR_NAME = "软件工程"
MAJOR_CODE = "080902SE"
GRADE = "2023"

# 6 门课：(课程代码, 课程名称, 课程性质, 学分)
courses = [
    ("SE101", "软件工程导论", "必修", 2.5),
    ("SE102", "数据结构与算法", "必修", 4.0),
    ("SE103", "操作系统", "必修", 4.0),
    ("SE104", "计算机网络", "必修", 3.5),
    ("SE105", "数据库原理", "必修", 3.5),
    ("SE106", "软件测试", "必修", 3.0),
]

# 30 名学生：学号 202301001..202301030，每班 5 人（按 k//5 分到 6 个班）
students = []
for k in range(30):
    sno = f"2023010{k+1:02d}"
    sname = f"测试学生{k+1:02d}"
    cls_idx = k // 5
    cls_name = f"软工2023-{courses[cls_idx][1]}班"
    students.append((k, sno, sname, cls_idx, cls_name))

# 每课 4 个考核点：(编号, 名称, 满分) —— 与数据库 assessment_point 实际数据一致（2026-06-30 修正）
points = [("AP1", "平时作业", 30), ("AP2", "实验报告", 40), ("AP3", "课程设计", 20), ("AP4", "期末考试", 10)]

def score(k, j, full):
    """确定性成绩：满分的 0.65~0.95，与 generate_test_data.py 公式一致"""
    ratio = 0.65 + 0.30 * ((k + j) % 10) / 10.0
    return round(full * ratio, 1)

bold = Font(bold=True)
wrap = Alignment(wrap_text=True, vertical="center", horizontal="center")

def style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = bold
        cell.alignment = wrap

# 1. 课程导入.xlsx（edu）
wb = openpyxl.Workbook(); ws = wb.active; ws.title = "课程"
ws.append(["所属专业*", "课程代码*", "课程名称*", "课程性质*", "学分*", "专业代码"])
for code, name, nat, cr in courses:
    ws.append([MAJOR_NAME, code, name, nat, cr, MAJOR_CODE])
style_header(ws, 6)
wb.save(os.path.join(OUT, "1-课程导入.xlsx"))

# 2. 学生导入.xlsx（edu）—— 学生导入不含班级，班级在绑定教学班时回填
wb = openpyxl.Workbook(); ws = wb.active; ws.title = "学生"
ws.append(["姓名*", "学号*", "年级", "专业代码*"])
for k, sno, sname, ci, cn in students:
    ws.append([sname, sno, GRADE, MAJOR_CODE])
style_header(ws, 4)
wb.save(os.path.join(OUT, "2-学生导入.xlsx"))

# 3. 教学班学生（6 个，每班 5 人）（edu，按教学班导入）
for ci, (code, name, nat, cr) in enumerate(courses):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "班级学生"
    ws.append(["学号*", "姓名"])
    for k, sno, sname, cidx, cn in students:
        if cidx == ci:
            ws.append([sno, sname])
    style_header(ws, 2)
    wb.save(os.path.join(OUT, f"3-教学班学生-{code}_{name}.xlsx"))

# 4. 成绩（6 个，每班 5 学生 × 4 考核点）（teacher，按教学班导入）
# 表头匹配后端动态模板：{编号}-{名称}\n满分:{满分}；导入按列序匹配 AP1..AP4
for ci, (code, name, nat, cr) in enumerate(courses):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "成绩录入"
    header = ["学号", "姓名"]
    for pc, pn, full in points:
        header.append(f"{pc}-{pn}\n满分:{full}")
    ws.append(header)
    for k, sno, sname, cidx, cn in students:
        if cidx == ci:
            row = [sno, sname]
            for j, (pc, pn, full) in enumerate(points):
                row.append(score(k, j, full))
            ws.append(row)
    style_header(ws, 2 + len(points))
    # 分数列居中
    for r in range(2, ws.max_row + 1):
        for c in range(3, 3 + len(points)):
            ws.cell(row=r, column=c).alignment = Alignment(horizontal="center")
    wb.save(os.path.join(OUT, f"4-成绩-{code}_{name}.xlsx"))

print("已生成 Excel：")
for f in sorted(os.listdir(OUT)):
    print("  ", f)
print(f"共 {len(os.listdir(OUT))} 个文件")
