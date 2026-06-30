# -*- coding: utf-8 -*-
"""
生成「内部贡献权重批量导入」测试 Excel。
- 正向：10-内部贡献权重导入.xlsx —— 从现有 weight_objective_indicator 导出，
  同一课程下支撑同一指标点的所有课程目标权重之和 = 1.0（合法，可一键导入通过校验）
- 反向：10-内部贡献权重导入-反向.xlsx —— 把某指标点的一个权重 +0.1，使该指标点和=1.1，
  用于验证导入校验拒绝 + 事务回滚
列顺序与后端 WeightObjectiveIndicatorExcel 一致：课程代码* / 课程目标编号* / 指标点编号* / 内部权重*
"""
import os
import subprocess
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

EXCEL_DIR = r"C:\Users\YU\Desktop\Computing_Platform_frontend\Computing_Platform\测试数据-课堂演示\excel"
HEADERS = ["课程代码*", "课程目标编号*", "指标点编号*", "内部权重*"]

MYSQL = [
    "mysql", "-h", "120.27.221.219", "-P", "3306", "-u", "yu", "-p125103Dbqwan",
    "--default-character-set=utf8mb4", "graduation_achievement", "-B", "-N", "-e",
    "SELECT c.course_code, co.obj_code, ip.indicator_code, w.inner_weight "
    "FROM weight_objective_indicator w "
    "JOIN course c ON c.id=w.course_id "
    "JOIN course_objective co ON co.id=w.objective_id "
    "JOIN indicator_point ip ON ip.id=w.indicator_id "
    "WHERE w.is_deleted=0 "
    "ORDER BY c.course_code, ip.indicator_code, co.obj_code",
]


def fetch_rows():
    out = subprocess.run(MYSQL, capture_output=True, text=True, encoding="utf-8")
    if out.returncode != 0:
        raise RuntimeError(out.stderr)
    lines = [ln for ln in out.stdout.strip().split("\n") if ln.strip()]
    return [ln.split("\t") for ln in lines]


def write_excel(path, data):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "内部贡献权重"
    ws.append(HEADERS)
    for row in data:
        ws.append(row)
    head_fill = PatternFill("solid", fgColor="D9E1F2")
    for c in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = head_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for r in range(2, ws.max_row + 1):
        for c in range(1, len(HEADERS) + 1):
            ws.cell(row=r, column=c).alignment = Alignment(horizontal="center", vertical="center")
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 12
    wb.save(path)


def main():
    rows = fetch_rows()
    print(f"从 weight_objective_indicator 导出 {len(rows)} 行")

    # 校验：同 (course, indicator) 权重和应=1.0
    from collections import defaultdict
    sums = defaultdict(float)
    for course_code, _, ind, w in rows:
        sums[(course_code, ind)] += float(w)
    bad = {k: round(v, 4) for k, v in sums.items() if abs(v - 1.0) > 0.0001}
    print(f"同(课程,指标点)权重合计不为1.0的有 {len(bad)} 组（应为0）:", dict(list(bad.items())[:3]))

    # 正向
    write_excel(os.path.join(EXCEL_DIR, "10-内部贡献权重导入.xlsx"), rows)
    print("OK 正向 -> 10-内部贡献权重导入.xlsx")

    # 反向：把第一行权重 +0.1，破坏其所在指标点的合计
    bad_rows = [r[:] for r in rows]
    first_course, _, first_ind, first_w = bad_rows[0]
    bad_rows[0][3] = str(round(float(first_w) + 0.1, 4))
    # 计算受影响指标点的新合计，便于提示
    new_sum = round(sum(float(r[3]) for r in bad_rows if r[0] == first_course and r[2] == first_ind), 4)
    write_excel(os.path.join(EXCEL_DIR, "10-内部贡献权重导入-反向.xlsx"), bad_rows)
    print(f"OK 反向 -> 10-内部贡献权重导入-反向.xlsx（{first_course} 指标点 {first_ind} 合计被改为 {new_sum}，应被拒绝）")


if __name__ == "__main__":
    main()
